from flask import Flask, render_template, request, redirect, session, send_file
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os, json, io

app = Flask(__name__)
app.secret_key = "supersecret"
FILE_NAME = "oblik_mayna.xlsx"
USERS_FILE = "users.json"

def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Майно"
        ws.append(["ID", "Назва", "Одиниця", "Кількість", "Примітка"])
        wb.save(FILE_NAME)

def init_users():
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, "w") as f:
            users = {"admin": generate_password_hash("admin")}
            json.dump(users, f)

def get_items():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))

def save_items(items):
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Назва", "Одиниця", "Кількість", "Примітка"])
    for item in items:
        ws.append(item)
    wb.save(FILE_NAME)

@app.route("/", methods=["GET", "POST"])
def index():
    if "user" not in session:
        return redirect("/login")
    items = get_items()
    query = request.args.get("q", "").lower()
    if query:
        items = [item for item in items if query in str(item[1]).lower()]
    return render_template("index.html", items=items, query=query)

@app.route("/add", methods=["GET", "POST"])
def add():
    if "user" not in session:
        return redirect("/login")
    if request.method == "POST":
        items = get_items()
        item_id = len(items) + 1
        new_item = [
            item_id,
            request.form["name"],
            request.form["unit"],
            int(request.form["quantity"]),
            request.form["note"]
        ]
        items.append(new_item)
        save_items(items)
        return redirect("/")
    return render_template("add.html")

@app.route("/edit/<int:item_id>", methods=["GET", "POST"])
def edit(item_id):
    if "user" not in session:
        return redirect("/login")
    items = get_items()
    item = next((i for i in items if i[0] == item_id), None)
    if not item:
        return "Item not found"
    if request.method == "POST":
        item[1] = request.form["name"]
        item[2] = request.form["unit"]
        item[3] = int(request.form["quantity"])
        item[4] = request.form["note"]
        save_items(items)
        return redirect("/")
    return render_template("edit.html", item=item)

@app.route("/delete/<int:item_id>")
def delete(item_id):
    if "user" not in session:
        return redirect("/login")
    items = [i for i in get_items() if i[0] != item_id]
    for idx, item in enumerate(items):
        item[0] = idx + 1
    save_items(items)
    return redirect("/")

@app.route("/export/pdf")
def export_pdf():
    if "user" not in session:
        return redirect("/login")
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    items = get_items()
    y = height - 40
    p.setFont("Helvetica", 12)
    p.drawString(30, y, "Облік майна зв'язку:")
    y -= 20
    for item in items:
        p.drawString(30, y, f"{item[0]}. {item[1]} | {item[2]} | {item[3]} | {item[4]}")
        y -= 20
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="mayno.pdf", mimetype="application/pdf")

@app.route("/export/excel")
def export_excel():
    if "user" not in session:
        return redirect("/login")
    return send_file(FILE_NAME, as_attachment=True)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        with open(USERS_FILE) as f:
            users = json.load(f)
        username = request.form["username"]
        password = request.form["password"]
        if username in users and check_password_hash(users[username], password):
            session["user"] = username
            return redirect("/")
        return "Невірний логін або пароль"
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect("/login")

if __name__ == "__main__":
    init_excel()
    init_users()
    app.run(debug=True)